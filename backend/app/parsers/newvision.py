"""
新视野供应商数据包：问卷定位与解析、眼底 FDT 处理约定（设计 V1.1.0 20260512）。
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from backend.app.core.errors import AppError
from backend.app.parsers.questionnaire import QuestionnaireParseResult, _cell_to_date_iso

logger = logging.getLogger(__name__)

IMPORT_PARSER_VERSION = "newvision-v1.1.0"
PACKAGE_DATA_SUBDIRS = frozenset({"生物测量", "眼底照相", "OCT"})


def _is_temp_xlsx_filename(name: str) -> bool:
    if name.startswith("~$"):
        return True
    return bool(re.match(r"^\.~.+\.xlsx$", name, re.IGNORECASE))


def _path_touches_package_data_subdir(rel: str) -> bool:
    for part in Path(rel.replace("\\", "/")).parts:
        if part in PACKAGE_DATA_SUBDIRS:
            return True
    return False


def select_newvision_questionnaire_xlsx(extracted_rels: list[str]) -> str:
    """根层有效问卷：排除数据子目录内 xlsx 与 Office 临时文件；同深度仅允许 1 份。"""
    cands: list[str] = []
    for r in extracted_rels:
        r = r.replace("\\", "/")
        if not r.lower().endswith(".xlsx"):
            continue
        if _is_temp_xlsx_filename(Path(r).name):
            continue
        if _path_touches_package_data_subdir(r):
            continue
        cands.append(r)
    if not cands:
        raise AppError(
            "压缩包内缺失有效问卷表格（根目录需 1 份 xlsx，且不在生物测量/眼底照相/OCT 目录内）。",
            "DATASET_IMPORT_TABLE_MISSING",
        )
    depths = [len(Path(p).parts) for p in cands]
    shallow = min(depths)
    top = [p for p in cands if len(Path(p).parts) == shallow]
    if len(top) != 1:
        raise AppError(
            f"根目录有效问卷 xlsx 必须为 1 份，当前命中 {len(top)} 份。",
            "DATASET_IMPORT_MULTIPLE_QUESTIONNAIRE",
            details={"candidates": top},
        )
    return top[0]


def _norm_header_token(s: str) -> str:
    t = str(s or "")
    t = t.replace("\u3000", " ").replace("：", ":").replace("．", ".")
    t = re.sub(r"\s+", "", t.strip().lower())
    return t


_NEWVISION_PID_HEADER = _norm_header_token("(2)    ID ：___")


def _find_pid_column_index(headers: list[str]) -> int:
    for i, h in enumerate(headers):
        if _norm_header_token(h) == _NEWVISION_PID_HEADER:
            logger.info("newvision questionnaire PID column header matched: %r", h)
            return i
    for i, h in enumerate(headers):
        n = _norm_header_token(h)
        if n in ("患者id", "病人id", "pid", "patientid", "受试者id", "subjectid"):
            logger.info("newvision questionnaire PID column (legacy header): %r", h)
            return i
        if "患者" in n and "id" in n:
            logger.info("newvision questionnaire PID column (legacy heuristic): %r", h)
            return i
    raise AppError(
        "无法在表格中定位新视野患者 ID 列（需 “(2)    ID ：___” 或兼容列）。",
        "DATASET_IMPORT_STRUCTURE_INVALID",
    )


def _find_survey_date_column_index(headers: list[str]) -> Optional[int]:
    for i, h in enumerate(headers):
        n = _norm_header_token(h)
        if n in ("调查日期", "随访日期", "检查日期"):
            return i
    for i, h in enumerate(headers):
        if "日期" in _norm_header_token(h):
            return i
    return None


def parse_newvision_questionnaire_xlsx_bytes(data: bytes) -> QuestionnaireParseResult:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl is required for questionnaire import") from e

    buf = io.BytesIO(data)
    wb = load_workbook(buf, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise AppError("问卷表格为空。", "DATASET_IMPORT_TABLE_MISSING")

        headers = [str(h) if h is not None else "" for h in header_row]
        if not any(headers):
            raise AppError("问卷表格表头无效。", "DATASET_IMPORT_TABLE_MISSING")
        pid_ix = _find_pid_column_index(headers)
        date_ix = _find_survey_date_column_index(headers)

        nonempty_headers = [(i, str(title)) for i, title in enumerate(headers) if str(title).strip()]
        header_list = [t for _, t in nonempty_headers]

        columns: list[dict[str, Any]] = []
        for i, title in enumerate(headers):
            if not str(title).strip():
                continue
            key = f"q_{i}"
            columns.append(
                {
                    "columnKey": key,
                    "columnTitle": str(title),
                    "dataType": "STRING",
                    "sourceType": "QUESTIONNAIRE",
                    "displayOrder": i,
                }
            )

        out_rows: list[dict[str, Any]] = []
        excel_row = 1
        for row in rows_iter:
            excel_row += 1
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            pid_val = row[pid_ix] if pid_ix < len(row) else None
            if isinstance(pid_val, float) and not isinstance(pid_val, bool):
                pid = str(int(pid_val)) if pid_val == int(pid_val) else str(pid_val).strip()
            elif isinstance(pid_val, int) and not isinstance(pid_val, bool):
                pid = str(pid_val)
            else:
                pid = str(pid_val).strip() if pid_val is not None else ""
            if not pid:
                raise AppError(
                    "表格存在数据行患者 ID 为空，请检查导入文件。",
                    "DATASET_IMPORT_PATIENT_ID_MISSING",
                )
            survey_date: Optional[str] = None
            raw_by_header: dict[str, Any] = {}
            if date_ix is None:
                survey_date = None
            else:
                date_cell = row[date_ix] if date_ix < len(row) else None
                if date_cell is None or str(date_cell).strip() == "":
                    survey_date = None
                else:
                    survey_date, ok = _cell_to_date_iso(date_cell)
                    if not ok:
                        raise AppError(
                            "表格存在无法识别的调查日期，请检查导入文件。",
                            "DATASET_IMPORT_DATE_PARSE_FAILED",
                        )

            raw_cells: dict[str, Any] = {}
            normalized_cells: dict[str, Any] = {}
            for i, title in enumerate(headers):
                if not str(title).strip():
                    continue
                key = f"q_{i}"
                cell = row[i] if i < len(row) else None
                raw_cells[key] = cell
                raw_by_header[str(title)] = cell
                if cell is None:
                    normalized_cells[key] = None
                elif isinstance(cell, datetime):
                    normalized_cells[key] = cell.isoformat(sep=" ", timespec="seconds")
                elif isinstance(cell, date):
                    normalized_cells[key] = cell.isoformat()
                else:
                    normalized_cells[key] = cell

            normalized_cells["patientId"] = pid
            normalized_cells["surveyDate"] = survey_date
            nv_payload = {
                "vendor": "newvision",
                "sourceType": "QUESTIONNAIRE",
                "parserVersion": IMPORT_PARSER_VERSION,
                "headers": header_list,
                "rowIndex": excel_row,
                "raw": raw_by_header,
                "normalized": {"pid": pid, "checkDate": survey_date},
            }
            out_rows.append(
                {
                    "patient_id": pid,
                    "survey_date": survey_date,
                    "raw_cells": raw_cells,
                    "normalized_cells": normalized_cells,
                    "newvision_import_payload": nv_payload,
                }
            )

        if not out_rows:
            raise AppError("问卷表格中无有效数据行。", "DATASET_IMPORT_TABLE_MISSING")
        return QuestionnaireParseResult(columns=columns, rows=out_rows)
    finally:
        wb.close()


def fundus_fdt_maybe_jpeg(raw: bytes) -> tuple[Optional[bytes], bool]:
    """新视野 ``.fdt`` 样例为 JPEG 裸流：校验 ``FF D8`` 后原样作为 JPG 字节。"""
    if len(raw) >= 2 and raw[0:2] == b"\xff\xd8":
        return raw, True
    return None, False
