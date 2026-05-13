"""
Questionnaire (.xlsx) import — design §3.3 问卷表格.

Column detection is heuristic so hospital templates can be adjusted via configuration later.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional

from backend.app.core.errors import AppError

_log = logging.getLogger(__name__)


@dataclass
class QuestionnaireParseResult:
    columns: list[dict[str, Any]]
    rows: list[dict[str, Any]]
    """Row dicts include ``patient_id``, ``survey_date``, ``raw_cells``, ``normalized_cells``."""


def _norm_header(s: str) -> str:
    return re.sub(r"\s+", "", str(s or "").strip().lower())


def _compact_newvision_pid_header(s: str) -> str:
    """Design §4.1: normalize to match ``(2)ID:___`` (full/half width space and colon stripped)."""
    t = str(s or "").strip().replace("\u3000", " ")
    t = re.sub(r"\s+", "", t)
    return t.replace("：", ":").lower()


def _is_newvision_pid_header(header: str) -> bool:
    c = _compact_newvision_pid_header(header)
    return bool(re.match(r"^\(2\)id[:_]", c))


def _find_pid_col(headers: list[str]) -> int:
    for i, h in enumerate(headers):
        if _is_newvision_pid_header(h):
            _log.info("questionnaire PID column matched Newvision header: %r", h)
            return i
    for i, h in enumerate(headers):
        n = _norm_header(h)
        if n in ("患者id", "病人id", "pid", "patientid", "受试者id", "subjectid"):
            _log.info("questionnaire PID column matched generic header: %r", h)
            return i
        if "患者" in n and "id" in n:
            _log.info("questionnaire PID column matched generic header: %r", h)
            return i
    raise AppError(
        "无法在表格中定位患者 ID 列，请检查导入文件。",
        "DATASET_IMPORT_STRUCTURE_INVALID",
    )


def _find_date_col(headers: list[str]) -> Optional[int]:
    for i, h in enumerate(headers):
        n = _norm_header(h)
        if n in ("调查日期", "随访日期", "检查日期", "日期", "surveydate", "visitdate"):
            return i
        if "日期" in n:
            return i
    return None


def _is_empty_survey_date_placeholder(val: Any) -> bool:
    """Excel placeholders that mean 「无检查日期」（设计验收：记入 DATE_EMPTY，不阻断）。"""
    if val is None or val == "":
        return True
    if not isinstance(val, str):
        return False
    s = val.strip().replace("\u3000", "")
    if not s:
        return True
    if s in ("(空)", "（空）", "空", "无"):
        return True
    sl = s.lower()
    if sl in ("n/a", "na", "null", "none"):
        return True
    return False


def _excel_serial_to_iso(val: float) -> str | None:
    """Map Excel OOXML serial date to ``YYYY-MM-DD`` (day granularity)."""
    try:
        sn = float(val)
    except (TypeError, ValueError):
        return None
    if not (200 < sn < 600_000):
        return None
    try:
        base = datetime(1899, 12, 30)
        d = base + timedelta(days=int(sn))
        return d.date().isoformat()
    except (OverflowError, ValueError):
        return None


def _cell_to_date_iso(val: Any) -> tuple[Optional[str], bool]:
    """
    Returns ``(iso_date, valid)``. For invalid non-empty values, ``valid`` is False.
    """
    if val is None or val == "":
        return None, True
    if isinstance(val, datetime):
        return val.date().isoformat(), True
    if isinstance(val, date):
        return val.isoformat(), True
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        iso = _excel_serial_to_iso(float(val))
        if iso:
            return iso, True
    if _is_empty_survey_date_placeholder(val):
        return None, True
    s = str(val).strip()
    if not s:
        return None, True
    if len(s) == 8 and s.isdigit():
        try:
            return datetime.strptime(s, "%Y%m%d").date().isoformat(), True
        except ValueError:
            pass
    norm = s[:10].replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d",):
        try:
            return datetime.strptime(norm[:10], fmt).date().isoformat(), True
        except ValueError:
            continue
    return None, False


def parse_questionnaire_xlsx_bytes(data: bytes) -> QuestionnaireParseResult:
    """Parse questionnaire from xlsx bytes (works with FTP-only storage backends)."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl is required for questionnaire import") from e

    buf = io.BytesIO(data)
    wb = load_workbook(buf, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise AppError("问卷表格为空。", "DATASET_IMPORT_TABLE_MISSING")

        headers = [str(h) if h is not None else "" for h in header_row]
        if not any(headers):
            raise AppError("问卷表格表头无效。", "DATASET_IMPORT_TABLE_MISSING")

        pid_ix = _find_pid_col(headers)
        date_ix = _find_date_col(headers)

        columns: list[dict[str, Any]] = []
        for i, title in enumerate(headers):
            if not str(title).strip():
                continue
            key = f"q_{i}"
            columns.append(
                {
                    "columnKey": key,
                    "columnTitle": str(title),
                    "dataType": "STRING",
                    "sourceType": "QUESTIONNAIRE",
                    "displayOrder": i,
                }
            )

        out_rows: list[dict[str, Any]] = []
        for row in rows_iter:
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            pid_val = row[pid_ix] if pid_ix < len(row) else None
            if isinstance(pid_val, float) and not isinstance(pid_val, bool):
                pid = str(int(pid_val)) if pid_val == int(pid_val) else str(pid_val).strip()
            elif isinstance(pid_val, int) and not isinstance(pid_val, bool):
                pid = str(pid_val)
            else:
                pid = str(pid_val).strip() if pid_val is not None else ""
            if not pid:
                raise AppError(
                    "表格存在数据缺失患者id，请检查导入文件。",
                    "DATASET_IMPORT_PATIENT_ID_MISSING",
                )
            survey_date: Optional[str] = None
            if date_ix is not None and date_ix < len(row):
                survey_date, ok = _cell_to_date_iso(row[date_ix])
                if not ok:
                    raise AppError(
                        "表格存在无法识别的调查日期，请检查导入文件。",
                        "DATASET_IMPORT_DATE_PARSE_FAILED",
                    )

            raw_cells: dict[str, Any] = {}
            normalized_cells: dict[str, Any] = {}
            for i, title in enumerate(headers):
                if not str(title).strip():
                    continue
                key = f"q_{i}"
                cell = row[i] if i < len(row) else None
                raw_cells[key] = cell
                if cell is None:
                    normalized_cells[key] = None
                elif isinstance(cell, datetime):
                    normalized_cells[key] = cell.isoformat(sep=" ", timespec="seconds")
                elif isinstance(cell, date):
                    normalized_cells[key] = cell.isoformat()
                else:
                    normalized_cells[key] = cell

            if date_ix is not None:
                normalized_cells["surveyDate"] = survey_date

            out_rows.append(
                {
                    "patient_id": pid,
                    "survey_date": survey_date,
                    "raw_cells": raw_cells,
                    "normalized_cells": normalized_cells,
                }
            )

        if not out_rows:
            raise AppError("问卷表格中无有效数据行。", "DATASET_IMPORT_TABLE_MISSING")

        return QuestionnaireParseResult(columns=columns, rows=out_rows)
    finally:
        wb.close()


def parse_questionnaire_xlsx(path: Path) -> QuestionnaireParseResult:
    return parse_questionnaire_xlsx_bytes(path.read_bytes())
