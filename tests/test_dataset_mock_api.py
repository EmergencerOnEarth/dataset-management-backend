from __future__ import annotations

import hashlib
import io
import json
import time
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from fastapi.testclient import TestClient

from backend.app.main import app
from backend.app.parsers.image_stubs import STUB_JPEG_BYTES


@pytest.fixture
def client():
    with TestClient(app) as c:
        yield c


def _minimal_zip_with_xlsx() -> bytes:
    xbio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["患者ID", "调查日期", "姓名"])
    ws.append(["LGTA00087", "2026-03-03", "测试用户"])
    wb.save(xbio)
    xbio.seek(0)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("问卷/演示.xlsx", xbio.read())
    return buf.getvalue()


def data(response):
    payload = response.json()
    assert payload["code"] == 0, payload
    assert payload["traceId"]
    return payload["data"]


def test_import_config_and_directory_list(client: TestClient):
    config = data(client.get("/api/v1/dataset-import/config"))
    assert config["maxFileSize"] == 8_388_608_000
    assert config["allowedExtensions"] == [".zip"]

    listing = data(client.get("/api/v1/dataset-directories", params={"pageNo": 1, "pageSize": 5}))
    assert listing["total"] >= 1
    for rec in listing["records"]:
        assert rec["canView"] == (rec["importStatus"] == "SUCCESS")


def test_upload_flow_and_import_task(client: TestClient):
    content = _minimal_zip_with_xlsx()
    file_hash = hashlib.sha256(content).hexdigest()
    check = data(
        client.post(
            "/api/v1/dataset-upload/instant-check",
            json={
                "fileName": "测试上传数据.zip",
                "fileSize": len(content),
                "fileHash": file_hash,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    assert check["hit"] is False

    upload = data(
        client.post(
            "/api/v1/dataset-upload/uploads",
            json={
                "fileName": "测试上传数据.zip",
                "fileSize": len(content),
                "fileHash": file_hash,
                "chunkSize": 1024 * 1024,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    assert upload["partCount"] == 1
    n = len(content)
    part = data(
        client.put(
            f"/api/v1/dataset-upload/uploads/{upload['uploadId']}/parts/1",
            content=content,
            headers={
                "X-Part-Hash": file_hash,
                "Content-Type": "application/octet-stream",
                "Content-Range": f"bytes 0-{n - 1}/{n}",
            },
        )
    )
    assert part["uploadedParts"] == [1]

    completed = data(
        client.post(
            f"/api/v1/dataset-upload/uploads/{upload['uploadId']}/complete",
            json={"fileHash": file_hash},
        )
    )
    assert completed["uploadStatus"] == "MERGED"

    created = data(
        client.post(
            "/api/v1/dataset-directories",
            json={
                "directoryName": "测试目录",
                "directoryDescription": "mock 接口测试",
                "fileId": completed["fileId"],
                "originalFileName": "测试上传数据.zip",
            },
        )
    )
    assert created["importStatus"] == "IMPORTING"

    tid = created["importTaskId"]
    st = None
    task = {}
    for _ in range(100):
        task = data(client.get(f"/api/v1/dataset-import/tasks/{tid}"))
        st = task["importStatus"]
        if st in ("SUCCESS", "FAILED"):
            break
        time.sleep(0.05)
    assert st == "SUCCESS", task
    assert task["progress"] == 100


def test_frame_url_helper():
    from types import SimpleNamespace

    from backend.app.services.directory_service import _frame_url

    oct_img = SimpleNamespace(
        image_id="img_oct",
        parsed_path=(
            "/dataset/import/parsed/dir_x/oct_frames/tag/"
            "od-3dscan-macular-001.frames/frame_00000.png"
        ),
    )
    assert _frame_url(oct_img) == "/api/v1/dataset-files/img_oct/frame/"

    jpg_only = SimpleNamespace(
        image_id="img_jpg",
        parsed_path="/dataset/import/parsed/dir_x/oct/foo.jpg",
    )
    assert _frame_url(jpg_only) is None

    no_parsed = SimpleNamespace(image_id="img_none", parsed_path=None)
    assert _frame_url(no_parsed) is None


def test_records_patient_images_and_exports(client: TestClient):
    records = data(client.get("/api/v1/dataset-directories/dir_demo_001/records"))
    assert records["columns"]
    assert records["records"][0]["patientId"] == "LGTA00087"

    timeline = data(client.get("/api/v1/dataset-directories/dir_demo_001/patients/LGTA00087/timeline"))
    assert timeline["dates"][0]["surveyDate"] == "2026-03-03"

    images = data(
        client.get(
            "/api/v1/dataset-directories/dir_demo_001/patients/LGTA00087/images",
            params={"surveyDate": "2026-03-03"},
        )
    )
    assert images["records"][0]["imageId"] == "img_001"

    detail = data(client.get("/api/v1/dataset-directories/dir_demo_001/patients/LGTA00087/images/img_001"))
    assert detail["metadata"]["width"] == 1024
    assert images["records"][0]["originalUrl"] == "/api/v1/dataset-files/img_001/original"
    assert detail["originalUrl"] == "/api/v1/dataset-files/img_001/original"
    assert images["records"][0].get("frameUrl") is None
    assert detail.get("frameUrl") is None

    oct_images = data(
        client.get(
            "/api/v1/dataset-directories/dir_demo_001/patients/LGTA00101/images",
            params={"surveyDate": "2026-03-06"},
        )
    )
    assert oct_images["records"][0]["imageId"] == "img_002"
    assert oct_images["records"][0].get("frameUrl") is None

    directory_export = data(
        client.post("/api/v1/dataset-directories/export", json={"directoryIds": ["dir_demo_001"]})
    )
    assert directory_export["exportType"] == "DATASET_DIRECTORY"

    patient_export = data(
        client.post("/api/v1/dataset-directories/dir_demo_001/patients/LGTA00087/export", json={})
    )
    assert patient_export["exportType"] == "DATASET_PATIENT"

    time.sleep(0.3)


def test_error_response_for_invalid_upload_type(client: TestClient):
    response = client.post(
        "/api/v1/dataset-upload/instant-check",
        json={
            "fileName": "bad.xlsx",
            "fileSize": 100,
            "fileHash": "abc",
            "businessType": "DATASET_IMPORT",
        },
    )
    assert response.status_code == 400
    payload = response.json()
    assert payload["errorCode"] == "DATASET_IMPORT_FILE_TYPE_INVALID"


def test_content_range_required_422(client: TestClient):
    tiny = _minimal_zip_with_xlsx()
    h = hashlib.sha256(tiny).hexdigest()
    u = data(
        client.post(
            "/api/v1/dataset-upload/uploads",
            json={
                "fileName": "t.zip",
                "fileSize": len(tiny),
                "fileHash": h,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    r = client.put(
        f"/api/v1/dataset-upload/uploads/{u['uploadId']}/parts/1",
        content=tiny,
        headers={"Content-Type": "application/octet-stream"},
    )
    assert r.status_code == 422
    assert r.json()["errorCode"] == "DATASET_VALIDATION_ERROR"


def test_content_range_vs_part_number_mismatch_400(client: TestClient):
    """验收 I-11b：第二片却声明文件首段区间。"""
    n = 1024 * 1024 + 1  # 大于一片，强制 partCount=2（最小分片 1MiB）
    blob = b"a" * n
    h = hashlib.sha256(blob).hexdigest()
    upl = data(
        client.post(
            "/api/v1/dataset-upload/uploads",
            json={
                "fileName": "twoparts.zip",
                "fileSize": len(blob),
                "fileHash": h,
                "chunkSize": 1024 * 1024,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    assert upl["partCount"] == 2
    uid = upl["uploadId"]
    client.put(
        f"/api/v1/dataset-upload/uploads/{uid}/parts/1",
        content=blob[:1024 * 1024],
        headers={
            "Content-Type": "application/octet-stream",
            "Content-Range": f"bytes 0-{1024 * 1024 - 1}/{n}",
        },
    )
    r = client.put(
        f"/api/v1/dataset-upload/uploads/{uid}/parts/2",
        content=blob[1024 * 1024 :],
        headers={
            "Content-Type": "application/octet-stream",
            # 末片长度为 1，但区间却声明为第一字节 → 序号与区间不一致（I-11b）
            "Content-Range": f"bytes 0-0/{n}",
        },
    )
    assert r.status_code == 400
    assert r.json()["errorCode"] == "DATASET_UPLOAD_RANGE_PART_MISMATCH"


def test_upload_after_merged_returns_409(client: TestClient):
    tiny = _minimal_zip_with_xlsx()
    fh = hashlib.sha256(tiny).hexdigest()
    u = data(
        client.post(
            "/api/v1/dataset-upload/uploads",
            json={
                "fileName": "once.zip",
                "fileSize": len(tiny),
                "fileHash": fh,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    n = len(tiny)
    client.put(
        f"/api/v1/dataset-upload/uploads/{u['uploadId']}/parts/1",
        content=tiny,
        headers={
            "Content-Type": "application/octet-stream",
            "Content-Range": f"bytes 0-{n - 1}/{n}",
        },
    )
    data(client.post(f"/api/v1/dataset-upload/uploads/{u['uploadId']}/complete", json={"fileHash": fh}))
    again = client.put(
        f"/api/v1/dataset-upload/uploads/{u['uploadId']}/parts/1",
        content=tiny,
        headers={
            "Content-Type": "application/octet-stream",
            "Content-Range": f"bytes 0-{n - 1}/{n}",
        },
    )
    assert again.status_code == 409


def test_import_pipeline_with_fdt_placeholder(client: TestClient):
    """含 .fdt：走眼底分支与 registry 解码桩，不因未定义变量失败。"""
    buf = io.BytesIO()
    xbio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["患者ID", "调查日期"])
    ws.append(["LGTA00087", "2026-03-03"])
    wb.save(xbio)
    xbio.seek(0)
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("问卷/眼底.xlsx", xbio.read())
        zf.writestr("fundus_placeholder.fdt", b"placeholder-fdt")
    content = buf.getvalue()
    fh = hashlib.sha256(content).hexdigest()
    u = data(
        client.post(
            "/api/v1/dataset-upload/uploads",
            json={
                "fileName": "with_fdt.zip",
                "fileSize": len(content),
                "fileHash": fh,
                "chunkSize": 1024 * 1024,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    client.put(
        f"/api/v1/dataset-upload/uploads/{u['uploadId']}/parts/1",
        content=content,
        headers={
            "Content-Type": "application/octet-stream",
            "Content-Range": f"bytes 0-{len(content) - 1}/{len(content)}",
        },
    )
    completed = data(
        client.post(f"/api/v1/dataset-upload/uploads/{u['uploadId']}/complete", json={"fileHash": fh})
    )
    created = data(
        client.post(
            "/api/v1/dataset-directories",
            json={
                "directoryName": "含fdt测",
                "fileId": completed["fileId"],
            },
        )
    )
    tid = created["importTaskId"]
    task = {}
    for _ in range(120):
        task = data(client.get(f"/api/v1/dataset-import/tasks/{tid}"))
        if task["importStatus"] in ("SUCCESS", "FAILED"):
            break
        time.sleep(0.05)
    assert task["importStatus"] == "SUCCESS"
    assert task.get("assetCount", 0) >= 1


def test_validation_422_missing_required_instant_check(client: TestClient):
    r = client.post(
        "/api/v1/dataset-upload/instant-check",
        json={"fileName": "a.zip", "fileSize": 10},
    )
    assert r.status_code == 422


def test_validation_422_missing_business_type_instant_check(client: TestClient):
    r = client.post(
        "/api/v1/dataset-upload/instant-check",
        json={
            "fileName": "a.zip",
            "fileSize": 10,
            "fileHash": "ab" * 32,
        },
    )
    assert r.status_code == 422


def test_page_no_zero_422(client: TestClient):
    r = client.get("/api/v1/dataset-directories", params={"pageNo": 0})
    assert r.status_code == 422


def test_patient_export_rejects_empty_patient(client: TestClient):
    r = client.post(
        "/api/v1/dataset-directories/dir_demo_001/patients/__no_one__/export",
        json={},
    )
    assert r.status_code == 400
    assert r.json()["errorCode"] == "DATASET_EXPORT_PATIENT_EMPTY"


def test_oct_sidecar_adds_dynamic_columns(client: TestClient):
    buf = io.BytesIO()
    xbio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["患者ID", "调查日期"])
    ws.append(["P_OCT_01", "2026-04-01"])
    wb.save(xbio)
    xbio.seek(0)
    sidecar = json.dumps({"eye": {"sphere": -2.5}}).encode()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("问卷/oct.xlsx", xbio.read())
        zf.writestr("检查/OCT/P_OCT_01/2026-04-01/scan.dat", b"oct-dat-bytes")
        zf.writestr("检查/OCT/P_OCT_01/2026-04-01/scan.json", sidecar)
    content = buf.getvalue()
    fh = hashlib.sha256(content).hexdigest()
    u = data(
        client.post(
            "/api/v1/dataset-upload/uploads",
            json={
                "fileName": "octzip.zip",
                "fileSize": len(content),
                "fileHash": fh,
                "chunkSize": 1024 * 1024,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    n = len(content)
    client.put(
        f"/api/v1/dataset-upload/uploads/{u['uploadId']}/parts/1",
        content=content,
        headers={
            "Content-Type": "application/octet-stream",
            "Content-Range": f"bytes 0-{n - 1}/{n}",
        },
    )
    completed = data(
        client.post(f"/api/v1/dataset-upload/uploads/{u['uploadId']}/complete", json={"fileHash": fh})
    )
    created = data(
        client.post(
            "/api/v1/dataset-directories",
            json={"directoryName": "OCT列测", "fileId": completed["fileId"]},
        )
    )
    tid = created["importTaskId"]
    did = created["directoryId"]
    task = {}
    for _ in range(120):
        task = data(client.get(f"/api/v1/dataset-import/tasks/{tid}"))
        if task["importStatus"] in ("SUCCESS", "FAILED"):
            break
        time.sleep(0.05)
    assert task["importStatus"] == "SUCCESS"
    recs = data(client.get(f"/api/v1/dataset-directories/{did}/records", params={"pageSize": 20}))
    col_keys = {c["columnKey"] for c in recs["columns"]}
    assert "oct_eye_sphere" in col_keys
    row0 = next(r for r in recs["records"] if r["patientId"] == "P_OCT_01")
    assert row0["cells"].get("oct_eye_sphere") == -2.5


def test_vendor_marker_zip_rejected_until_implemented(client: TestClient):
    """未注册的供应商特征文件应使导入失败（layout 非 SUPPORTED）"""
    buf = io.BytesIO()
    xbio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["患者ID", "调查日期"])
    ws.append(["P_X", "2026-05-01"])
    wb.save(xbio)
    xbio.seek(0)
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(".dataset_vendor/acme_v1.marker", b"1")
        zf.writestr("问卷/x_vendor.xlsx", xbio.read())
    content = buf.getvalue()
    fh = hashlib.sha256(content).hexdigest()
    u = data(
        client.post(
            "/api/v1/dataset-upload/uploads",
            json={
                "fileName": "acme.zip",
                "fileSize": len(content),
                "fileHash": fh,
                "chunkSize": 1024 * 1024,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    n = len(content)
    client.put(
        f"/api/v1/dataset-upload/uploads/{u['uploadId']}/parts/1",
        content=content,
        headers={
            "Content-Type": "application/octet-stream",
            "Content-Range": f"bytes 0-{n - 1}/{n}",
        },
    )
    completed = data(
        client.post(f"/api/v1/dataset-upload/uploads/{u['uploadId']}/complete", json={"fileHash": fh})
    )
    created = data(
        client.post(
            "/api/v1/dataset-directories",
            json={"directoryName": "Vendor测", "fileId": completed["fileId"]},
        )
    )
    tid = created["importTaskId"]
    task = {}
    for _ in range(120):
        task = data(client.get(f"/api/v1/dataset-import/tasks/{tid}"))
        if task["importStatus"] in ("SUCCESS", "FAILED"):
            break
        time.sleep(0.05)
    assert task["importStatus"] == "FAILED"


def test_oct_path_folder_case_insensitive(client: TestClient):
    """ISSUE-07：路径中为 ``oct``/``Oct`` 等目录名时需识别 OCT .dat"""
    buf = io.BytesIO()
    xbio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["患者ID", "调查日期"])
    ws.append(["P_OCT_LC", "2026-05-10"])
    wb.save(xbio)
    xbio.seek(0)
    sidecar = json.dumps({"layer": {"thickness": 3.14}}).encode()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("问卷/oct_lc.xlsx", xbio.read())
        zf.writestr("vendor/oct/P_OCT_LC/2026-05-10/scan_lc.dat", b"oct-d")
        zf.writestr("vendor/oct/P_OCT_LC/2026-05-10/scan_lc.json", sidecar)
    content = buf.getvalue()
    fh = hashlib.sha256(content).hexdigest()
    u = data(
        client.post(
            "/api/v1/dataset-upload/uploads",
            json={
                "fileName": "oct_lc.zip",
                "fileSize": len(content),
                "fileHash": fh,
                "chunkSize": 1024 * 1024,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    n = len(content)
    client.put(
        f"/api/v1/dataset-upload/uploads/{u['uploadId']}/parts/1",
        content=content,
        headers={
            "Content-Type": "application/octet-stream",
            "Content-Range": f"bytes 0-{n - 1}/{n}",
        },
    )
    completed = data(
        client.post(f"/api/v1/dataset-upload/uploads/{u['uploadId']}/complete", json={"fileHash": fh})
    )
    created = data(
        client.post(
            "/api/v1/dataset-directories",
            json={"directoryName": "OCT小写目录", "fileId": completed["fileId"]},
        )
    )
    tid = created["importTaskId"]
    task = {}
    for _ in range(120):
        task = data(client.get(f"/api/v1/dataset-import/tasks/{tid}"))
        if task["importStatus"] in ("SUCCESS", "FAILED"):
            break
        time.sleep(0.05)
    assert task["importStatus"] == "SUCCESS"


def test_patient_export_contains_parsed_jpeg_and_oct_json(client: TestClient):
    """患者导出含眼底解析图（JPEG）；OCT sidecar 仅当有 *-001.dat 入库影像时随 PARSED_OCT_DAT 导出。"""
    from backend.app.core.config import get_settings
    from backend.app.db.models import ExportRecord
    from backend.app.db.session import get_session_factory
    from backend.app.storage.backend import get_storage

    buf = io.BytesIO()
    xbio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["患者ID", "调查日期"])
    ws.append(["P_EXP01", "2026-06-01"])
    wb.save(xbio)
    xbio.seek(0)
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("问卷/pexp.xlsx", xbio.read())
        zf.writestr("眼底照相/P_EXP01/2026-06-01/f1.fdt", STUB_JPEG_BYTES)
    content = buf.getvalue()
    fh = hashlib.sha256(content).hexdigest()
    u = data(
        client.post(
            "/api/v1/dataset-upload/uploads",
            json={
                "fileName": "mix_export.zip",
                "fileSize": len(content),
                "fileHash": fh,
                "chunkSize": 1024 * 1024,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    n = len(content)
    client.put(
        f"/api/v1/dataset-upload/uploads/{u['uploadId']}/parts/1",
        content=content,
        headers={
            "Content-Type": "application/octet-stream",
            "Content-Range": f"bytes 0-{n - 1}/{n}",
        },
    )
    completed = data(
        client.post(f"/api/v1/dataset-upload/uploads/{u['uploadId']}/complete", json={"fileHash": fh})
    )
    created = data(
        client.post(
            "/api/v1/dataset-directories",
            json={"directoryName": "导出混测", "fileId": completed["fileId"]},
        )
    )
    tid = created["importTaskId"]
    did = created["directoryId"]
    for _ in range(150):
        task = data(client.get(f"/api/v1/dataset-import/tasks/{tid}"))
        if task["importStatus"] in ("SUCCESS", "FAILED"):
            break
        time.sleep(0.05)
    assert task["importStatus"] == "SUCCESS"

    exp_resp = data(
        client.post(
            f"/api/v1/dataset-directories/{did}/patients/P_EXP01/export",
            json={"includeParsedImages": True, "includeOriginalAttachments": True},
        )
    )
    export_id = exp_resp["exportRecordId"]

    ftp_path_final: str | None = None
    for _ in range(180):
        fac = get_session_factory()
        db = fac()
        try:
            er = db.get(ExportRecord, export_id)
            if er and er.export_status in ("DONE", "FAILED"):
                assert er.export_status == "DONE", (er.export_status, er.failure_reason)
                ftp_path_final = er.ftp_path
                assert ftp_path_final
                break
        finally:
            db.close()
        time.sleep(0.05)
    assert ftp_path_final is not None

    st = get_storage(get_settings())
    zblob = zipfile.ZipFile(io.BytesIO(st.get_bytes(ftp_path_final)))
    names = zblob.namelist()
    parsed_members = [n for n in names if n.startswith("images/parsed/") and n.lower().endswith(".jpg")]
    assert parsed_members, names
    assert zblob.read(parsed_members[0]).startswith(b"\xff\xd8\xff")
    assert any(n.startswith("images/") and n.lower().endswith(".fdt") for n in names)


def test_export_list_detail_and_download(client: TestClient):
    """API-18/19：导出任务列表、详情与受控下载。"""
    from backend.app.db.models import ExportRecord
    from backend.app.db.session import get_session_factory

    content = _minimal_zip_with_xlsx()
    task, _tid = _flow_upload_import(client, content, "导出查询测")
    assert task["importStatus"] == "SUCCESS"
    did = task["directoryId"]

    dir_exp = data(
        client.post(
            "/api/v1/dataset-directories/export",
            json={"directoryIds": [did], "includeParsedImages": True},
        )
    )
    pat_exp = data(
        client.post(
            f"/api/v1/dataset-directories/{did}/patients/LGTA00087/export",
            json={"includeParsedImages": True},
        )
    )

    listing = data(client.get("/api/v1/dataset-exports", params={"offset": 0, "limit": 20}))
    assert listing["total"] >= 2
    assert listing["offset"] == 0
    assert listing["limit"] == 20
    types = {r["exportType"] for r in listing["records"]}
    assert "DATASET_DIRECTORY" in types
    assert "DATASET_PATIENT" in types
    for rec in listing["records"]:
        assert rec["exportRecordId"]
        assert rec["exportTypeName"]
        assert rec["exportStatusName"]
        assert "downloadable" in rec
        assert "summary" in rec
        assert "ftp" not in (rec.get("downloadUrl") or "").lower()

    page2 = data(client.get("/api/v1/dataset-exports", params={"offset": 1, "limit": 1}))
    assert page2["total"] == listing["total"]
    assert page2["offset"] == 1
    assert len(page2["records"]) == 1

    filtered = data(
        client.get(
            "/api/v1/dataset-exports",
            params={"exportType": "DATASET_PATIENT", "exportStatus": "PREPARING"},
        )
    )
    assert all(r["exportType"] == "DATASET_PATIENT" for r in filtered["records"])

    detail = data(client.get(f"/api/v1/dataset-exports/{dir_exp['exportRecordId']}"))
    assert detail["exportRecordId"] == dir_exp["exportRecordId"]
    assert detail["exportType"] == "DATASET_DIRECTORY"
    assert detail["payload"]["directoryIds"] == [did]
    assert detail["downloadable"] is False
    assert detail["downloadUrl"] is None

    missing = client.get("/api/v1/dataset-exports/exp_not_exists")
    assert missing.status_code == 404

    bad_page = client.get("/api/v1/dataset-exports", params={"offset": 0, "limit": 0})
    assert bad_page.status_code == 422

    done_id = pat_exp["exportRecordId"]
    for _ in range(180):
        fac = get_session_factory()
        db = fac()
        try:
            er = db.get(ExportRecord, done_id)
            if er and er.export_status in ("DONE", "FAILED"):
                assert er.export_status == "DONE", (er.export_status, er.failure_reason)
                break
        finally:
            db.close()
        time.sleep(0.05)

    done_detail = data(client.get(f"/api/v1/dataset-exports/{done_id}"))
    assert done_detail["exportStatus"] == "DONE"
    assert done_detail["downloadable"] is True
    assert done_detail["downloadUrl"] == f"/api/v1/dataset-exports/{done_id}/download"
    assert "ftp" not in done_detail["downloadUrl"].lower()

    dl = client.get(done_detail["downloadUrl"])
    assert dl.status_code == 200
    assert dl.headers.get("content-type", "").startswith("application/zip")
    assert zipfile.is_zipfile(io.BytesIO(dl.content))


def _flow_upload_import(client: TestClient, content: bytes, directory_name: str, *, complete_extra: dict | None = None) -> tuple[dict, str]:
    fh = hashlib.sha256(content).hexdigest()
    u = data(
        client.post(
            "/api/v1/dataset-upload/uploads",
            json={
                "fileName": "selftest.zip",
                "fileSize": len(content),
                "fileHash": fh,
                "chunkSize": 1024 * 1024,
                "businessType": "DATASET_IMPORT",
            },
        )
    )
    n = len(content)
    client.put(
        f"/api/v1/dataset-upload/uploads/{u['uploadId']}/parts/1",
        content=content,
        headers={
            "Content-Type": "application/octet-stream",
            "Content-Range": f"bytes 0-{n - 1}/{n}",
        },
    )
    payload = {"fileHash": fh}
    if complete_extra:
        payload.update(complete_extra)
    completed = data(
        client.post(f"/api/v1/dataset-upload/uploads/{u['uploadId']}/complete", json=payload)
    )
    created = data(
        client.post(
            "/api/v1/dataset-directories",
            json={"directoryName": directory_name, "fileId": completed["fileId"]},
        )
    )
    tid = created["importTaskId"]
    task = {}
    for _ in range(150):
        task = data(client.get(f"/api/v1/dataset-import/tasks/{tid}"))
        if task["importStatus"] in ("SUCCESS", "FAILED"):
            break
        time.sleep(0.05)
    return task, tid


def test_complete_upload_accepts_parts_as_etag_strings(client: TestClient):
    """05-11 联调 / TC-API-07：complete 携带 etag 字符串数组不得 422。"""
    content = _minimal_zip_with_xlsx()
    task, _ = _flow_upload_import(
        client,
        content,
        "合并入参测",
        complete_extra={"parts": ["5496bccb4581bd9c", "aabbdeadbeef"]},
    )
    assert task["importStatus"] == "SUCCESS"


def test_tc_d001_import_without_questionnaire_fails(client: TestClient):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("readme.txt", b"no xlsx")
    task, _ = _flow_upload_import(client, buf.getvalue(), "无问卷测")
    assert task["importStatus"] == "FAILED"
    assert "表格" in (task.get("failureReason") or "")


def test_tc_d002_import_empty_pid_fails(client: TestClient):
    xbio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["患者ID", "调查日期"])
    ws.append(["", "2026-04-01"])
    wb.save(xbio)
    xbio.seek(0)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("根/表.xlsx", xbio.read())
    task, _ = _flow_upload_import(client, buf.getvalue(), "空PID测")
    assert task["importStatus"] == "FAILED"
    fr = task.get("failureReason") or ""
    assert "患者" in fr or "ID" in fr


def test_tc_d003_import_invalid_survey_date_fails(client: TestClient):
    xbio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["患者ID", "调查日期"])
    ws.append(["P1", "not-a-real-date"])
    wb.save(xbio)
    xbio.seek(0)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("根/表.xlsx", xbio.read())
    task, _ = _flow_upload_import(client, buf.getvalue(), "日期非法测")
    assert task["importStatus"] == "FAILED"
    assert "日期" in (task.get("failureReason") or "")


def test_dc005_two_root_questionnaires_fails(client: TestClient):
    """两份根层问卷 xlsx → DATASET_IMPORT_MULTIPLE_QUESTIONNAIRE。"""
    def _xlsx_row(pid: str):
        bio = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["患者ID", "调查日期"])
        ws.append([pid, "2026-05-01"])
        wb.save(bio)
        bio.seek(0)
        return bio.read()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("问卷/a.xlsx", _xlsx_row("A1"))
        zf.writestr("问卷/b.xlsx", _xlsx_row("B1"))
    task, _ = _flow_upload_import(client, buf.getvalue(), "双问卷测")
    assert task["importStatus"] == "FAILED"
    assert "1 份" in (task.get("failureReason") or "") or "问卷" in (task.get("failureReason") or "")


def test_tc_sec001_zip_slip_rejected(client: TestClient):
    xbio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["患者ID", "调查日期"])
    ws.append(["P1", "2026-05-01"])
    wb.save(xbio)
    xbio.seek(0)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("问卷/q.xlsx", xbio.read())
        zf.writestr("../evil.txt", b"bad")
    task, _ = _flow_upload_import(client, buf.getvalue(), "ZipSlip测")
    assert task["importStatus"] == "FAILED"
    assert "安全" in (task.get("failureReason") or "") or "路径" in (task.get("failureReason") or "")


def test_tc_sec005_gbk_zip_filename_decoded():
    """SEC-05: ZipInfo 无 UTF-8 flag、文件名为 cp437 存储的 GBK 内容时应还原中文。"""
    from unittest.mock import MagicMock
    from backend.app.services.import_pipeline import _decode_zip_member_name

    # Simulate a Windows-created zip: GBK bytes read back as cp437 mojibake
    gbk_bytes = "问卷/问卷.xlsx".encode("gbk")
    mojibake = gbk_bytes.decode("cp437")

    m = MagicMock(spec=zipfile.ZipInfo)
    m.flag_bits = 0  # no UTF-8 flag
    m.filename = mojibake

    decoded = _decode_zip_member_name(m)
    assert "问卷" in decoded, f"中文目录名未正确还原: {decoded!r}"

    # If UTF-8 flag is set, should pass through unchanged
    m2 = MagicMock(spec=zipfile.ZipInfo)
    m2.flag_bits = 0x0800
    m2.filename = "问卷/问卷.xlsx"
    assert _decode_zip_member_name(m2) == "问卷/问卷.xlsx"


def test_tc_nv001_newvision_header_columns_import(client: TestClient):
    """NV-01：根问卷 ``(2)    ID ：___`` + ``调查日期``；含 2026/3/3 白名单日期。"""
    xbio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["(2)    ID ：___", "调查日期", "备注"])
    ws.append(["LGTA_NV01", "2026/3/3", "x"])
    wb.save(xbio)
    xbio.seek(0)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("包根/2026年问卷.xlsx", xbio.read())
    task, _ = _flow_upload_import(client, buf.getvalue(), "NV表头测")
    assert task["importStatus"] == "SUCCESS"
    # directory id from last created - need to get from API - poll doesn't return id
    listing = data(client.get("/api/v1/dataset-directories", params={"pageNo": 1, "pageSize": 20}))
    did = next(r["directoryId"] for r in listing["records"] if r["directoryName"] == "NV表头测")
    recs = data(client.get(f"/api/v1/dataset-directories/{did}/records", params={"pageSize": 20}))
    row0 = recs["records"][0]
    assert row0["patientId"] == "LGTA_NV01"
    assert row0["surveyDate"] == "2026-03-03"
    payload = row0["cells"].get("newvisionImportPayload")
    assert payload is not None
    assert payload["normalized"]["pid"] == "LGTA_NV01"
    assert payload["normalized"]["checkDate"] == "2026-03-03"


def test_nv10_oct_dat_header_golden_file():
    """NV-10 / DC-08：样例 dat 前 1024 字节可解析为 EOD header。"""
    from backend.app.parsers.newvision_oct import HEADER_SIZE, parse_oct_header

    repo = Path(__file__).resolve().parents[1]
    dat_path = (
        repo / "test-data/upload-samples/local/测试上传数据/中航数据/2026/OCT/2026-3-3/"
        "X08-data(2026-03-03-2026-03-03)/database/info-data/50/LGTA00087/"
        "x08-rds/20260303/od-3dscan-macular-20260303-092714-001.dat"
    )
    if not dat_path.is_file():
        pytest.skip(f"missing golden sample: {dat_path}")
    raw = dat_path.read_bytes()[:HEADER_SIZE]
    assert len(raw) == HEADER_SIZE
    hdr = parse_oct_header(raw, prefer_time_t_8=True)
    assert hdr.signature.strip("\x00") == "EOD"
    assert hdr.scan.nFrames >= 1