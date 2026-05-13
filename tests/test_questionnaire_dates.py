from __future__ import annotations

from backend.app.parsers.questionnaire import _cell_to_date_iso


def test_yyyymmdd_string():
    assert _cell_to_date_iso("20260303") == ("2026-03-03", True)


def test_excel_serial_reasonable_range():
    iso, ok = _cell_to_date_iso(45000)
    assert ok and iso is not None and len(iso) == 10


def test_newvision_fixed_pid_header():
    from io import BytesIO

    from openpyxl import Workbook

    from backend.app.parsers.questionnaire import parse_questionnaire_xlsx_bytes

    buf = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["(2)    ID ：___", "调查日期", "备注"])
    ws.append(["NV_PID_01", "2026-03-03", "note"])
    wb.save(buf)
    r = parse_questionnaire_xlsx_bytes(buf.getvalue())
    assert len(r.rows) == 1
    assert r.rows[0]["patient_id"] == "NV_PID_01"
    assert r.rows[0]["survey_date"] == "2026-03-03"


def test_empty_survey_date_placeholder_kong():
    assert _cell_to_date_iso("(空)") == (None, True)
    assert _cell_to_date_iso("（空）") == (None, True)
    assert _cell_to_date_iso("NA") == (None, True)


def test_questionnaire_two_rows_placeholder_and_real_date():
    from io import BytesIO

    from openpyxl import Workbook

    from backend.app.parsers.questionnaire import parse_questionnaire_xlsx_bytes

    buf = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["(2)    ID ：___", "调查日期"])
    ws.append(["P_A", "(空)"])
    ws.append(["P_B", "2026-03-02"])
    wb.save(buf)
    r = parse_questionnaire_xlsx_bytes(buf.getvalue())
    assert len(r.rows) == 2
    assert r.rows[0]["survey_date"] is None and r.rows[0]["patient_id"] == "P_A"
    assert r.rows[1]["survey_date"] == "2026-03-02"


def test_newvision_pid_header_halfwidth_colon():
    from io import BytesIO

    from openpyxl import Workbook

    from backend.app.parsers.questionnaire import parse_questionnaire_xlsx_bytes

    buf = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["(2) ID : ___", "调查日期"])
    ws.append([42, "20260303"])
    wb.save(buf)
    r = parse_questionnaire_xlsx_bytes(buf.getvalue())
    assert r.rows[0]["patient_id"] == "42"
